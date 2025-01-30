"""
Module: display
This module contains logic to display the coffee stock in a tabular format.
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from prettytable import PrettyTable
from coffee_management.data_structure import fetch_stock

def display_coffee_stock():
    """
    Displays the coffee inventory in a table format using PrettyTable and writes it to an Excel sheet.
    """
    coffee_stock = fetch_stock()
    if not coffee_stock:
        print("from display_coffee_stock: Coffee stock is empty. Please initialize the stock.")
        return

    print(f"{'-' * 6} Below is the Inventory for the Coffee packs in Inventory {'-' * 6} \n")

    table = PrettyTable()
    table.field_names = ["Coffee Pack", "Size", "Type", "Price (€)", "Quantity"]

    data = []
    for name, sizes in coffee_stock.items():
        for size, types in sizes.items():
            for coffee_type, details in types.items():
                row = [name, size, coffee_type, f"{details['price']:.2f}", details["quantity"]]
                table.add_row(row)
                data.append(row)

    print(table)

    # Load the existing workbook or create a new one if it doesn't exist
    try:
        wb = load_workbook("coffee_stock_inventory.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Coffee Stock Inventory"
        # Add headers if creating a new workbook
        headers = ["Coffee Pack", "Size", "Type", "Price (€)", "Quantity"]
        ws.append(headers)

    # Clear existing data while preserving the header
    ws.delete_rows(2, ws.max_row)

    # Add new data rows
    for row in data:
        ws.append(row)

    # Apply conditional formatting
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        quantity = row[4].value
        if quantity == 0:
            for cell in row:
                cell.fill = red_fill
        elif quantity <= 2:
            for cell in row:
                cell.fill = yellow_fill

    # Save the workbook
    wb.save("coffee_stock_inventory.xlsx")
    print("Inventory has been written to 'coffee_stock_inventory.xlsx'")
